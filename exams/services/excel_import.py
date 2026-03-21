import json
import unicodedata
from datetime import datetime

import openpyxl


COLUMN_MAPPINGS = {
    "nom": "last_name",
    "lastname": "last_name",
    "prenom": "first_name",
    "prénom": "first_name",
    "sexe": "gender",
    "genre": "gender",
    "date de naissance": "birth_date",
    "naissance": "birth_date",
}


def normalize(text: str) -> str:
    return unicodedata.normalize("NFD", text.lower().strip()).encode("ascii", "ignore").decode()


def parse_gender(value):
    if not value:
        return ""
    s = str(value).upper().strip()
    if s in ("M", "MASCULIN", "HOMME", "H"):
        return "M"
    if s in ("F", "FEMININ", "FÉMININ", "FEMME"):
        return "F"
    return ""


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_excel_file(file) -> dict:
    """
    Parse un fichier Excel uploadé et retourne les données des élèves.
    Retourne : {'success': bool, 'students': list, 'errors': list}
    """
    errors = []
    students = []

    try:
        wb = openpyxl.load_workbook(file, read_only=True)
    except Exception as e:
        return {"success": False, "students": [], "errors": [f"Erreur lecture fichier: {e}"]}

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return {"success": False, "students": [], "errors": ["Fichier vide ou sans données."]}

    headers = rows[0]
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            key = normalize(str(h))
            for mapping_key, mapping_val in COLUMN_MAPPINGS.items():
                if key == normalize(mapping_key):
                    col_map[i] = mapping_val
                    break

    mapped_fields = set(col_map.values())
    if "last_name" not in mapped_fields or "first_name" not in mapped_fields:
        return {"success": False, "students": [], "errors": ["Colonnes Nom/Prénom manquantes."]}

    for row_num, row in enumerate(rows[1:], start=2):
        student = {}
        for col_idx, field in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            if field == "gender":
                val = parse_gender(val)
            elif field == "birth_date":
                val = parse_date(val)
            else:
                val = str(val).strip() if val else ""
            student[field] = val

        if not student.get("last_name"):
            errors.append(f"Ligne {row_num}: Nom manquant")
            continue
        if not student.get("first_name"):
            errors.append(f"Ligne {row_num}: Prénom manquant")
            continue

        students.append(student)

    wb.close()
    return {"success": len(students) > 0, "students": students, "errors": errors}


# ──────────────────────────────────────────────
# Import complet d'examen (multi-onglets)
# ──────────────────────────────────────────────

SCHOOL_COLUMNS = {"nom": "name", "name": "name", "code": "code"}
SUBJECT_COLUMNS = {
    "nom": "name", "name": "name", "matiere": "name", "matière": "name",
    "coefficient": "coefficient", "coef": "coefficient",
    "note maximale": "max_score", "note max": "max_score", "max": "max_score",
}


def _find_sheet(wb, candidates: list[str]):
    """Trouve une feuille par nom (insensible à la casse)."""
    for name in wb.sheetnames:
        if normalize(name) in [normalize(c) for c in candidates]:
            return wb[name]
    return None


def parse_exam_excel(file) -> dict:
    """
    Importe un examen complet depuis un fichier Excel.

    Onglets attendus (noms flexibles) :
    - Établissements / Ecoles / Schools
    - Épreuves / Matières / Subjects
    - Élèves / Candidats / Students

    Retourne : {
        'success': bool,
        'schools': list[dict],
        'subjects': list[dict],
        'students': list[dict],
        'errors': list[str],
    }
    """
    result = {"success": False, "schools": [], "subjects": [], "students": [], "errors": []}

    try:
        wb = openpyxl.load_workbook(file, read_only=True)
    except Exception as e:
        result["errors"].append(f"Erreur lecture fichier : {e}")
        return result

    # ── Établissements ──
    ws_schools = _find_sheet(wb, ["etablissements", "écoles", "ecoles", "schools"])
    if ws_schools:
        rows = list(ws_schools.iter_rows(values_only=True))
        if len(rows) >= 2:
            headers = rows[0]
            col_map = {}
            for i, h in enumerate(headers or []):
                if h:
                    key = normalize(str(h))
                    for mk, mv in SCHOOL_COLUMNS.items():
                        if key == normalize(mk):
                            col_map[i] = mv
                            break
            for row_num, row in enumerate(rows[1:], start=2):
                school = {}
                for ci, field in col_map.items():
                    val = row[ci] if ci < len(row) else None
                    school[field] = str(val).strip() if val else ""
                if school.get("name"):
                    result["schools"].append(school)
                else:
                    result["errors"].append(f"Établissements ligne {row_num} : nom manquant")

    # ── Épreuves ──
    ws_subjects = _find_sheet(wb, ["epreuves", "épreuves", "matieres", "matières", "subjects"])
    if ws_subjects:
        rows = list(ws_subjects.iter_rows(values_only=True))
        if len(rows) >= 2:
            headers = rows[0]
            col_map = {}
            for i, h in enumerate(headers or []):
                if h:
                    key = normalize(str(h))
                    for mk, mv in SUBJECT_COLUMNS.items():
                        if key == normalize(mk):
                            col_map[i] = mv
                            break
            for row_num, row in enumerate(rows[1:], start=2):
                subject = {}
                for ci, field in col_map.items():
                    val = row[ci] if ci < len(row) else None
                    if field in ("coefficient", "max_score"):
                        try:
                            subject[field] = float(val) if val else None
                        except (ValueError, TypeError):
                            subject[field] = None
                    else:
                        subject[field] = str(val).strip() if val else ""
                if subject.get("name"):
                    result["subjects"].append(subject)
                else:
                    result["errors"].append(f"Épreuves ligne {row_num} : nom manquant")

    # ── Élèves ──
    ws_students = _find_sheet(wb, ["eleves", "élèves", "candidats", "students"])
    if ws_students:
        rows = list(ws_students.iter_rows(values_only=True))
        if len(rows) >= 2:
            # Réutilise la logique existante
            headers = rows[0]
            # Ajouter mapping établissement
            extended_mappings = {**COLUMN_MAPPINGS, "etablissement": "school_name", "école": "school_name", "ecole": "school_name", "school": "school_name"}
            col_map = {}
            for i, h in enumerate(headers or []):
                if h:
                    key = normalize(str(h))
                    for mk, mv in extended_mappings.items():
                        if key == normalize(mk):
                            col_map[i] = mv
                            break
            for row_num, row in enumerate(rows[1:], start=2):
                student = {}
                for ci, field in col_map.items():
                    val = row[ci] if ci < len(row) else None
                    if field == "gender":
                        val = parse_gender(val)
                    elif field == "birth_date":
                        val = parse_date(val)
                    else:
                        val = str(val).strip() if val else ""
                    student[field] = val
                if student.get("last_name") and student.get("first_name"):
                    result["students"].append(student)
                else:
                    result["errors"].append(f"Élèves ligne {row_num} : nom ou prénom manquant")

    if not result["schools"] and not result["subjects"] and not result["students"]:
        result["errors"].append("Aucun onglet reconnu (Établissements, Épreuves, Élèves).")
    else:
        result["success"] = True

    wb.close()
    return result


# ──────────────────────────────────────────────
# Import d'examen depuis JSON
# ──────────────────────────────────────────────

def parse_exam_json(file) -> dict:
    """
    Importe un examen complet depuis un fichier JSON.

    Supporte deux formats :
    1. Format backup (ancienne app React/Tauri) :
       { "version": "1.0", "data": { "exam": {...}, "schools": {"schools": [...]}, ... } }
    2. Format simple :
       { "schools": [...], "subjects": [...], "students": [...] }
    """
    result = {
        "success": False, "schools": [], "subjects": [], "students": [],
        "errors": [], "exam_meta": None,
    }

    try:
        raw = file.read()
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8")
        data = json.loads(raw)
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        result["errors"].append(f"Erreur lecture JSON : {e}")
        return result

    if not isinstance(data, dict):
        result["errors"].append("Le fichier JSON doit contenir un objet.")
        return result

    # ── Détection du format backup (ancienne app) ──
    if "data" in data and isinstance(data["data"], dict):
        return _parse_backup_json(data, result)

    # ── Format simple ──
    return _parse_simple_json(data, result)


def _find_key(d, candidates):
    """Trouve une clé dans un dict de manière flexible (insensible casse/accents)."""
    for c in candidates:
        if c in d:
            return d[c]
    for c in candidates:
        nc = normalize(c)
        for k in d:
            if normalize(k) == nc:
                return d[k]
    return None


def _parse_backup_json(data, result):
    """Parse le format backup de l'ancienne app React/Tauri/Prisma."""
    inner = data["data"]

    # ── Métadonnées examen ──
    exam_raw = inner.get("exam", {})
    if exam_raw:
        result["exam_meta"] = {
            "name": exam_raw.get("examName", ""),
            "year": exam_raw.get("examYear"),
            "passing_grade": exam_raw.get("passingGrade"),
            "max_grade": exam_raw.get("maxGrade"),
            "is_locked": exam_raw.get("status") == "locked",
        }

    # ── Établissements (data.schools.schools) ──
    schools_container = inner.get("schools", {})
    raw_schools = schools_container.get("schools", []) if isinstance(schools_container, dict) else []
    school_id_map = {}  # old id -> school name
    for s in raw_schools:
        if isinstance(s, dict) and s.get("name"):
            name = str(s["name"]).strip()
            code = str(s.get("code") or "").strip()
            result["schools"].append({"name": name, "code": code})
            school_id_map[s.get("id")] = name

    # ── Épreuves (data.subjects.subjects) ──
    subjects_container = inner.get("subjects", {})
    raw_subjects = subjects_container.get("subjects", []) if isinstance(subjects_container, dict) else []
    for s in raw_subjects:
        if isinstance(s, dict) and s.get("name"):
            coef = s.get("coefficient")
            max_s = s.get("maxScore") or s.get("max_score")
            try:
                coef = float(coef) if coef else None
            except (ValueError, TypeError):
                coef = None
            try:
                max_s = float(max_s) if max_s else None
            except (ValueError, TypeError):
                max_s = None
            result["subjects"].append({
                "name": str(s["name"]).strip(),
                "coefficient": coef,
                "max_score": max_s,
            })

    # ── Élèves (data.students.students) avec schoolId -> nom ──
    students_container = inner.get("students", {})
    raw_students = students_container.get("students", []) if isinstance(students_container, dict) else []
    for i, s in enumerate(raw_students, start=1):
        if not isinstance(s, dict):
            result["errors"].append(f"Élève #{i} : format invalide")
            continue
        last_name = s.get("lastName", "") or ""
        first_name = s.get("firstName", "") or ""
        gender = parse_gender(s.get("gender"))
        birth = s.get("birthDate")
        birth_date = parse_date(birth) if birth else None
        candidate_number = s.get("candidateNumber", "")
        is_absent = bool(s.get("isAbsent", False))
        school_id = s.get("schoolId")
        school_name = school_id_map.get(school_id, "")

        if last_name and first_name:
            result["students"].append({
                "last_name": str(last_name).strip(),
                "first_name": str(first_name).strip(),
                "gender": gender,
                "birth_date": birth_date,
                "school_name": school_name,
                "candidate_number": str(candidate_number).strip(),
                "is_absent": is_absent,
            })
        else:
            result["errors"].append(f"Élève #{i} : nom ou prénom manquant")

    # ── Scores (data.scores) ──
    raw_scores = inner.get("scores")
    if raw_scores and isinstance(raw_scores, list):
        result["scores"] = raw_scores

    if not result["schools"] and not result["subjects"] and not result["students"]:
        result["errors"].append("Aucune donnée trouvée dans le backup.")
    else:
        result["success"] = True

    return result


def _parse_simple_json(data, result):
    """Parse le format JSON simple."""

    # ── Établissements ──
    raw_schools = _find_key(data, ["schools", "etablissements", "écoles", "ecoles"]) or []
    for i, s in enumerate(raw_schools, start=1):
        if isinstance(s, str):
            result["schools"].append({"name": s, "code": ""})
        elif isinstance(s, dict):
            name = _find_key(s, ["name", "nom"]) or ""
            code = _find_key(s, ["code"]) or ""
            if name:
                result["schools"].append({"name": str(name).strip(), "code": str(code).strip()})
            else:
                result["errors"].append(f"Établissement #{i} : nom manquant")

    # ── Épreuves ──
    raw_subjects = _find_key(data, ["subjects", "epreuves", "épreuves", "matieres", "matières"]) or []
    for i, s in enumerate(raw_subjects, start=1):
        if isinstance(s, str):
            result["subjects"].append({"name": s, "coefficient": 1, "max_score": 20})
        elif isinstance(s, dict):
            name = _find_key(s, ["name", "nom", "matiere", "matière"]) or ""
            coef = _find_key(s, ["coefficient", "coef"])
            max_s = _find_key(s, ["max_score", "note_max", "max", "maxScore"])
            try:
                coef = float(coef) if coef else None
            except (ValueError, TypeError):
                coef = None
            try:
                max_s = float(max_s) if max_s else None
            except (ValueError, TypeError):
                max_s = None
            if name:
                result["subjects"].append({"name": str(name).strip(), "coefficient": coef, "max_score": max_s})
            else:
                result["errors"].append(f"Épreuve #{i} : nom manquant")

    # ── Élèves ──
    raw_students = _find_key(data, ["students", "eleves", "élèves", "candidats"]) or []
    for i, s in enumerate(raw_students, start=1):
        if not isinstance(s, dict):
            result["errors"].append(f"Élève #{i} : format invalide")
            continue
        last_name = _find_key(s, ["last_name", "lastName", "nom", "lastname"]) or ""
        first_name = _find_key(s, ["first_name", "firstName", "prenom", "prénom", "firstname"]) or ""
        gender = parse_gender(_find_key(s, ["gender", "sexe", "genre"]))
        birth = _find_key(s, ["birth_date", "birthDate", "naissance", "date_naissance"])
        school_name = _find_key(s, ["school_name", "etablissement", "école", "ecole", "school"]) or ""

        birth_date = parse_date(birth) if birth else None

        if last_name and first_name:
            result["students"].append({
                "last_name": str(last_name).strip(),
                "first_name": str(first_name).strip(),
                "gender": gender,
                "birth_date": birth_date,
                "school_name": str(school_name).strip(),
            })
        else:
            result["errors"].append(f"Élève #{i} : nom ou prénom manquant")

    if not result["schools"] and not result["subjects"] and not result["students"]:
        result["errors"].append("Aucune donnée trouvée (schools, subjects, students).")
    else:
        result["success"] = True

    return result
