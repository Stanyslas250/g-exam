import re
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def generate_results_excel(exam, students_data: list[dict]) -> BytesIO:
    """
    Génère un fichier Excel des résultats d'un examen.
    Retourne un BytesIO contenant le fichier .xlsx
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Résultats"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Résultats - {exam.name} ({exam.year})"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["N°", "N° Candidat", "Nom", "Prénom", "Moyenne", "Rang", "Mention"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for idx, s in enumerate(students_data, 1):
        row = idx + 3
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=s.get("candidate_number", "")).border = thin_border
        ws.cell(row=row, column=3, value=s.get("last_name", "")).border = thin_border
        ws.cell(row=row, column=4, value=s.get("first_name", "")).border = thin_border

        avg_cell = ws.cell(row=row, column=5, value=s.get("average"))
        avg_cell.number_format = "0.00"
        avg_cell.border = thin_border

        ws.cell(row=row, column=6, value=s.get("rank", "")).border = thin_border
        ws.cell(row=row, column=7, value=s.get("mention", "")).border = thin_border

    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 15

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_results_by_school_excel(exam, students_data: list[dict]) -> BytesIO:
    """Un onglet par établissement avec les résultats des élèves."""
    from collections import defaultdict

    wb = openpyxl.Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    by_school: dict[str, list] = defaultdict(list)
    for s in students_data:
        by_school[s.get("school_name", "—")].append(s)

    def _safe_title(name: str) -> str:
        t = re.sub(r"[\[\]*?:/\\]", "_", name).strip() or "Ecole"
        return t[:31]

    for school_name in sorted(by_school.keys()):
        ws = wb.create_sheet(title=_safe_title(school_name))
        ws.merge_cells("A1:G1")
        t = ws["A1"]
        t.value = f"{school_name} — {exam.name} ({exam.year})"
        t.font = Font(bold=True, size=12)
        t.alignment = Alignment(horizontal="center")

        headers = ["N°", "N° Candidat", "Nom", "Prénom", "Moyenne", "Rang", "Mention"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        rows = sorted(by_school[school_name], key=lambda x: (x.get("rank") or 9999, x.get("last_name", "")))
        for idx, s in enumerate(rows, 1):
            row = idx + 3
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=s.get("candidate_number", "")).border = thin_border
            ws.cell(row=row, column=3, value=s.get("last_name", "")).border = thin_border
            ws.cell(row=row, column=4, value=s.get("first_name", "")).border = thin_border
            avg_cell = ws.cell(row=row, column=5, value=s.get("average"))
            avg_cell.number_format = "0.00"
            avg_cell.border = thin_border
            ws.cell(row=row, column=6, value=s.get("rank", "")).border = thin_border
            ws.cell(row=row, column=7, value=s.get("mention", "")).border = thin_border

        for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col_letter].width = 15

    if not wb.sheetnames:
        ws = wb.create_sheet("Vide")
        ws["A1"] = "Aucune donnée"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
